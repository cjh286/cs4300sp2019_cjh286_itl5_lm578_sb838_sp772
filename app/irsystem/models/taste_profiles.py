from collections import defaultdict 
import csv
import re
import openpyxl
import os
import json

def tokenize(text): 
    """ Taken from class code A1
    Returns a list of words that make up the text.
    
    Note: for simplicity, lowercase everything.
    Requirement: Use Regex to satisfy this function
    
    Params: {text: String}
    Returns: List 
    """
    # YOUR CODE HERE
    x = re.findall(r'\b([a-zA-Z]+)\b',text)
    for y in range(len(x)):
        char = x[y].lower()
        x[y] = char
    return x

# def create_flavor_dict():
#     flavor_dict = {}
#     script_path = os.path.abspath(__file__) 
#     path_list = script_path.split(os.sep)
#     script_directory = path_list[0:len(path_list)-4]
#     rel_path = 'scraped_data/reviews.csv'
#     path = "/".join(script_directory) + "/" + rel_path
#     with open(path, encoding="utf8") as csv_file:
#         csv_reader = csv.reader(csv_file, delimiter=',')
#         for row in csv_reader:
#             if not alc_dict.get(row[13]):
#                 alc_dict[row[13]] = tokenize(row[23])
#             else:
#                 alc_dict[row[13]].extend(tokenize(row[23]))



# def overlap():
#     liquor_store_output_drinks = []
#     script_path = os.path.abspath(__file__) 
#     path_list = script_path.split(os.sep)
#     script_directory = path_list[0:len(path_list)-4]
#     rel_path = 'scraped_data/liquor_store_output.txt'
#     path = "/".join(script_directory) + "/" + rel_path
#     with open(path, encoding="utf8") as json_file:
#         data = json.load(json_file)
#         for p in data['liquor_info']:
#             tokenized_drink_name = tokenize(p['name'])
#             liquor_store_output_drinks.append(tokenized_drink_name)

#     return liquor_store_output_drinks


def create_flavor_dict():
    ingred_dict = {} #store alc names and review text
    script_path = os.path.abspath(__file__) 
    path_list = script_path.split(os.sep)
    script_directory = path_list[0:len(path_list)-4]
    rel_path = 'scraped_data/ingredient_tastes.csv'
    path = "/".join(script_directory) + "/" + rel_path
    with open(path, encoding="utf8") as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0
        for row in csv_reader:
            if not ingred_dict.get(row[0]):
                ingred_dict[row[0]] = tokenize(row[3])
            else:
                ingred_dict[row[0]].extend(tokenize(row[3]))
        for row in csv_reader:
            if not ingred_dict.get(row[0]):
                ingred_dict[row[0]] = tokenize(row[4])
            else:
                ingred_dict[row[0]].extend(tokenize(row[4]))


    # # read in excel data as set
    tastes = []

    script_path = os.path.abspath(__file__) 
    path_list = script_path.split(os.sep)
    script_directory = path_list[0:len(path_list)-4]
    rel_path = 'scraped_data/taste_words.xlsx'
    path = "/".join(script_directory) + "/" + rel_path
    book = openpyxl.load_workbook(path)
    sheet = book.get_sheet_by_name('Sheet1')
    for i in range(1, sheet.max_row+1):
        tastes.append(sheet.cell(row=i, column=1).value)

    tasteset = set(tastes)


    flavors_dict = {}
    for ingredient in ingred_dict.keys():
        list_to_grab_negation = ingred_dict[ingredient]
    # # generic_alcohols = ["brandy", "gin", "rum", "schnapps", "tequila", "vodka", "whisky", "bitters"]
    # for drink in ingred_dict.keys():
    #     list_to_grab_negation = ingred_dict[drink]
        

        to_remove_from_set = []
        get_negative_phrase = {}
        for w in range(len(list_to_grab_negation)-1):
            if list_to_grab_negation[w] == 'no' or list_to_grab_negation[w] == 'not' or list_to_grab_negation[w] == 'never':
                to_remove_from_set.append(list_to_grab_negation[w+1])
                get_negative_phrase[list_to_grab_negation[w+1]] = list_to_grab_negation[w] + " " + list_to_grab_negation[w+1]

        negative_terms = set(to_remove_from_set)

        reviewset = set(ingred_dict.get(ingredient))
        for word in negative_terms:
            reviewset.remove(word)

        overlapping_flavors = reviewset.intersection(tasteset)
        #check if any of the negative terms like 'flavorful' from 'not flavorful' are in the tasteset
        overlapping_negative_flavors = negative_terms.intersection(tasteset)


        if len(list(overlapping_flavors)) != 0:
            flavors_dict[ingredient] = list(overlapping_flavors) #list of positive descriptive words
            for word in overlapping_negative_flavors: 
                flavors_dict[ingredient].append(get_negative_phrase[word]) #word is 'flavorful' so add 'not flavorful' to the dictionary

    return flavors_dict

def main():
    flavor_dict = create_flavor_dict()
    print(flavor_dict)
    # print(overlap())

if __name__ == "__main__":
    main()